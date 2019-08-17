import numpy, os, string, sys, copy, random

try:
    from openpyxl.workbook import Workbook
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.reader.excel import load_workbook
    from openpyxl.cell import get_column_letter
    from openpyxl.reader.excel import load_workbook
    haveOpenpyxl=True
except ImportError:
    haveOpenpyxl=False

class SingleStair(object):

    def __init__(self, name='', startVal=5, iniStepSize=16.8, phi=0.75, extraInfo=None, nTrials=5, minVal=0):
        self.name = name
        self.startVal = random.choice([startVal, -startVal])
        self.iniStepSize = iniStepSize
        self.stepSizeCurrent = iniStepSize
        self.stepSizes = []
        self.extraInfo = extraInfo
        self.phi = phi
        self.nReversals = 0
        self.Reversals = []
        
        self.nTrials = nTrials
        self.finished = False
        self.thisTrialN = -1
        self.otherData = {}
        self.data = []
        self.stimuli = []
        self.directions = []
        self.reversalStimuli = []
        self.currentDirection = 'start'
        self.correctCounter = 0
        self._nextStimuli = self.startVal
        self.minVal = minVal
        self.initialRule = 0

    def __iter__(self):
        return self

    def addResponse(self, result, stimuli=None):
        self.data.append(result)

        if stimuli!=None:
            self.stimuli.pop() # removes and returns the last item in the list
            self.stimuli.append(stimuli)

        if result == 1:
            if len(self.data) > 0 and self.data[-1] == result:
                self.correctCounter += 1
            else:
                self.correctCounter = 1
                self.currentDirection = 'down'
        else:
            if len(self.data) > 0 and self.data[-1] == result:
                self.correctCounter -= 1
            else:
                self.correctCounter = -1
                self.currentDirection = 'up'

        self.calculateNextStimuli()

    def addOtherData(self, dataName, value):
        if not dataName in self.otherData:
            if self.thisTrialN > 0:
                self.otherData[dataName] = [None] * (self.thisTrialN - 1)
            else:
                self.otherData[dataName] = []

        self.otherData[dataName].append(value)

    def calculateNextStimuli(self):
        #if len(self.reversalStimuli) < 1: # "<1" in the original file, not sure. I think it should be ">1".
        if self.data[-1] == 1:
            if self.currentDirection == 'up':
                reversal = True
            else:
                reversal = False
            self.currentDirection = 'down'
        else:
            if self.currentDirection == 'down':
                reversal = True
            else:
                reversal = False
            self.currentDirection = 'up'

        self.directions.append(self.currentDirection)

        if reversal:
            self.reversalStimuli.append(self.stimuli[-1])
            self.nReversals += 1
        else:
            self.reversalStimuli.append(None)
        self.Reversals.append(self.nReversals)

        # Test if we're done
        if self.thisTrialN > self.nTrials: 
            self.finished = True

        # Apply new step size
        if self.thisTrialN < 2:
            self._calcM(self.data[-1])
        else:
            self._calcKesten(self.data[-1])

    def next(self):
        if not self.finished:
            for key in self.otherData.keys():
                while len(self.otherData[key]) < self.thisTrialN:
                    self.otherData[key].append(None)
                    
            self.thisTrialN += 1
            self.stimuli.append(self._nextStimuli)
            return self._nextStimuli
        else:
            raise StopIteration

    def _calcM(self, result):
        self.stepSizeCurrent = self.iniStepSize * (result - self.phi)/len(self.stimuli)
        tmp_stimuli = abs(self.stimuli[-1]) - self.stepSizeCurrent
        self._nextStimuli = random.choice([tmp_stimuli, -tmp_stimuli])
        self.stepSizes.append(self.stepSizeCurrent)

    def _calcKesten(self, result):
    	# if self.stimuli[-1] == 0:
    	# 	result = 0
        self.stepSizeCurrent = self.iniStepSize * (result - self.phi)/(2 + self.nReversals)
        if result == 1:
        	if abs(self.stimuli[-1]) < self.stepSizeCurrent and self.stimuli[-1] != 0: # if the step size gets bigger than the current stimuli intensity, make the next simuli zero
        		self.stepSizeCurrent = abs(self.stimuli[-1])

        tmp_stimuli = abs(self.stimuli[-1]) - self.stepSizeCurrent
        self._nextStimuli = random.choice([tmp_stimuli, -tmp_stimuli])
        self.stepSizes.append(self.stepSizeCurrent)

    def saveAsExcel(self, fileName, sheetName='data', matrixOnly=False, appendFile=True):
        if not haveOpenpyxl:
            raise ImportError('openpyxl is required for saving files in Excel (xlsx) format, but was not found.')

        if not fileName.endswith('.xlsx'):
            fileName+='.xlsx'
        #create or load the file
        if appendFile and os.path.isfile(fileName):
            wb = load_workbook(fileName)
            newWorkbook=False
        else:
            wb = Workbook()#create new workbook
            newWorkbook=True

        ew = ExcelWriter(workbook = wb)

        if newWorkbook:
            ws = wb.worksheets[0]
            ws.title=sheetName
        else:
            ws=wb.create_sheet()
            ws.title=sheetName

        #write the data
        # trials data
        ws.cell('A1').value = 'Trial No'
        ws.cell('B1').value = 'Stimulis'
        ws.cell('C1').value = 'Step Size'
        ws.cell('D1').value = 'Response'
        for orienN, stimuli in enumerate(self.stimuli):
            rowN = orienN + 2
            ws.cell(column=1, row=rowN).value = orienN + 1
            ws.cell(column=2, row=rowN).value = stimuli
            ws.cell(column=3, row=rowN).value = self.stepSizes[orienN]
            ws.cell(column=4, row=rowN).value = self.data[orienN]

        # reversals data
        ws.cell('E1').value = 'Directions'
        ws.cell('F1').value = 'Reversal Numbers'
        for revN, revIntens in enumerate(self.directions):
            rowN = revN + 2
            ws.cell(column=5, row=rowN).value = self.directions[revN]
            ws.cell(column=6, row=rowN).value = self.Reversals[revN]

        # other data
        col = 1
        for key in self.otherData.keys():
            ws.cell(column= col + 6, row=1).value = key
            for rtN, rtIntens in enumerate(self.otherData[key]):
                rowN = rtN + 2
                ws.cell(column=col+6, row=rowN).value = rtIntens
            col+=1

        ew.save(filename = fileName)

class SimpleStair(object):
    def __init__(self, name='simple', startVal=10, nTrials=13):
        self.name = name
        self.startVal = startVal
        
        self.nTrials = nTrials
        self.finished = False
        self.thisTrialN = -1
        self.otherData = {}
        self.data = []
        self.stimuli = []
        self.correctCounter = 0
        self._nextStimuli = self.startVal

    def __iter__(self):
        return self

    def addResponse(self, result, stimuli=None):
        self.data.append(result)

        if stimuli!=None:
            self.stimuli.pop() # removes and returns the last item in the list
            self.stimuli.append(stimuli)

        if result == 1:
            if len(self.data) > 0 and self.data[-1] == result:
                self.correctCounter += 1
            else:
                self.correctCounter = 1
        else:
            if len(self.data) > 0 and self.data[-1] == result:
                self.correctCounter -= 1
            else:
                self.correctCounter = -1

        self.calculateNextStimuli()

    def addOtherData(self, dataName, value):
        if not dataName in self.otherData:
            if self.thisTrialN > 0:
                self.otherData[dataName] = [None] * (self.thisTrialN - 1)
            else:
                self.otherData[dataName] = []

        self.otherData[dataName].append(value)

    def calculateNextStimuli(self):

        # Test if we're done
        if self.thisTrialN > self.nTrials:
            self.finished = True

        # Apply new step size
        valPool = [self.startVal, -self.startVal]
        self._nextStimuli = random.choice(valPool)

    def next(self):
        if not self.finished:
            for key in self.otherData.keys():
                while len(self.otherData[key]) < self.thisTrialN:
                    self.otherData[key].append(None)
                    
            self.thisTrialN += 1
            self.stimuli.append(self._nextStimuli)
            return self._nextStimuli
        else:
            raise StopIteration

    def saveAsExcel(self, fileName, sheetName='data', matrixOnly=False, appendFile=True):
        if not haveOpenpyxl:
            raise ImportError('openpyxl is required for saving files in Excel (xlsx) format, but was not found.')

        if not fileName.endswith('.xlsx'):
            fileName+='.xlsx'
        #create or load the file
        if appendFile and os.path.isfile(fileName):
            wb = load_workbook(fileName)
            newWorkbook=False
        else:
            wb = Workbook()#create new workbook
            newWorkbook=True

        ew = ExcelWriter(workbook = wb)

        if newWorkbook:
            ws = wb.worksheets[0]
            ws.title=sheetName
        else:
            ws=wb.create_sheet()
            ws.title=sheetName

        #write the data
        # trials data
        ws.cell('A1').value = 'Trial No'
        ws.cell('B1').value = 'Stimulis'
        ws.cell('C1').value = 'Step Size'
        ws.cell('D1').value = 'Response'
        for orienN, stimuli in enumerate(self.stimuli):
            rowN = orienN + 2
            ws.cell(column=1, row=rowN).value = orienN + 1
            ws.cell(column=2, row=rowN).value = stimuli
            ws.cell(column=3, row=rowN).value = 0
            ws.cell(column=4, row=rowN).value = self.data[orienN]

        # other data
        col = 1
        for key in self.otherData.keys():
            ws.cell(column= col + 4, row=1).value = key
            for rtN, rtIntens in enumerate(self.otherData[key]):
                rowN = rtN + 2
                ws.cell(column=col+4, row=rowN).value = rtIntens

        ew.save(filename = fileName)

class MultiStairs(object):
    def __init__(self, conditions= None, nTrials=5, withSimple=True):
        
        self.conditions = conditions
        self.nTrials = nTrials
        self.finished = False
        self.totalTrials = 0

        self.withSimple = withSimple

        self.staircases = []
        self.runningStaircases = []
        self.thisPassRemaining = []
        self._createStairs()

        self._startNewPass()
        self.currentStaircase = self.thisPassRemaining[0]
        self._nextStimuli = self.currentStaircase._nextStimuli


    def _createStairs(self):
        defaults = {'iniStepSize': 10, 'nTrials': 10, 'extraInfo': None, 'minVal': 0.2} # iniStepSize = 16.8; 8.4; 4.2

        for condition in self.conditions:
            startVal = condition['startVal']

            for paramName in defaults:
                if paramName in condition.keys():
                    val = condition[paramName]
                else:
                    val = defaults[paramName]

                exec('%s=%s' %(paramName, repr(val)))

            if not (self.nTrials == defaults['nTrials']):
                nTrials = self.nTrials

            thisStair = SingleStair(startVal, iniStepSize=iniStepSize, nTrials=nTrials, extraInfo=extraInfo, minVal=minVal)
            thisStair.condition = condition

            self.staircases.append(thisStair)
            self.runningStaircases.append(thisStair)

        if self.withSimple:
            thisStair = SimpleStair()
            thisStair.condition = {'label': 'Simple'}
            self.staircases.append(thisStair)
            self.runningStaircases.append(thisStair)

    def __iter__(self):
        return self 

    def next(self):
        #create a new set for this pass if needed
        if not hasattr(self, 'thisPassRemaining') or self.thisPassRemaining==[]:
            if len(self.runningStaircases)>0:
                self._startNewPass()
            else:
                self.finished=True
                raise StopIteration

        self.currentStaircase = self.thisPassRemaining.pop(0)
        self._nextStimuli = self.currentStaircase.next()

        if not self.finished:
            return self._nextStimuli, self.currentStaircase.condition
        else:
            raise StopIteration

    def _startNewPass(self):
        self.thisPassRemaining = copy.copy(self.runningStaircases)
        numpy.random.shuffle(self.thisPassRemaining)

    def addResponse(self, result, stimuli=None):
        self.currentStaircase.addResponse(result, stimuli)
        if self.currentStaircase.finished:
            self.runningStaircases.remove(self.currentStaircase)

        self.totalTrials += 1

    def addOtherData(self, name, value):
        self.currentStaircase.addOtherData(name, value)

    def saveAsExcel(self, fileName, matrixOnly=False, appendFile=False):
        append = appendFile
        for thisStair in self.staircases:
            #make a filename
            label = thisStair.condition['label']
            thisStair.saveAsExcel(fileName=fileName, sheetName=label, matrixOnly=matrixOnly, appendFile=append)
            append = True

